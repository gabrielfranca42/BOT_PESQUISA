import requests
from bs4 import BeautifulSoup
import nltk
from nltk.tokenize import sent_tokenize
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def extrair_texto(url, timeout=10):
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=timeout)
        r.raise_for_status()

        soup = BeautifulSoup(r.text, "html.parser")

        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()

        texto = " ".join(p.get_text() for p in soup.find_all("p"))
        return texto.strip()

    except Exception as e:
        print(f"Erro ao ler {url}: {e}")
        return ""

def resumir_texto(texto, max_frases=5):
    frases = sent_tokenize(texto, language="portuguese")
    return " ".join(frases[:max_frases])

def ajustar_formatacao_excel(arquivo):
    wb = load_workbook(arquivo)
    ws = wb.active

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Estilo do cabeçalho
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustes de colunas e células
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)

            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

            # Quebra de linha e alinhamento para o resumo
            if col_letter == "B" and row > 1:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

        # Larguras personalizadas
        if col_letter == "A":
            ws.column_dimensions[col_letter].width = 60
        elif col_letter == "B":
            ws.column_dimensions[col_letter].width = 120
        else:
            ws.column_dimensions[col_letter].width = min(max_length + 2, 80)

    wb.save(arquivo)

def processar_links(
    arquivo_links="links_filtrados.txt",
    arquivo_saida="resumos.xlsx"
):
    with open(arquivo_links, "r", encoding="utf-8") as f:
        links = [l.strip() for l in f if l.strip()]

    dados = []

    for i, link in enumerate(links, 1):
        print(f"\nLendo ({i}/{len(links)}): {link}")

        texto = extrair_texto(link)

        if len(texto) < 300:
            print("Conteúdo muito curto, pulando...")
            continue

        resumo = resumir_texto(texto)

        dados.append({
            "Link": link,
            "Resumo": resumo
        })

    df = pd.DataFrame(dados)
    df.to_excel(arquivo_saida, index=False)

    ajustar_formatacao_excel(arquivo_saida)

    print(f"\n✅ Excel formatado e salvo em {arquivo_saida}")

if __name__ == "__main__":
    processar_links()